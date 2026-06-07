"""
04_03_generate_own_report.py — Consolidated own_dataset domain-adaptation report.

Pulls together the findings produced by the Stage 02/03 analyses into one
visual Excel (own_dataset_report.xlsx). Reads only the saved JSON + PNG of the
existing analyses, so it needs no GPU and regenerates instantly whenever those
analyses are re-run with more photos.

Inputs:
  run_outputs/_finetune/*.json               (+ matching .png)
  run_outputs/_per_fruit/per_fruit.json      (+ per_fruit.png)
  run_outputs/_error_analysis/error_analysis.json (+ error_analysis.png)
  run_outputs/_ensemble/*.json               (+ ensemble_top3_mean.png)

Output: own_dataset_report.xlsx
  Sheets: Summary | Fine-tuning | Per-fruit | Hardest images | Ensemble

Usage:
    python 04_03_generate_own_report.py
"""

import sys
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

RUN = Path("run_outputs")
OUT = Path("own_dataset_report.xlsx")

FT_DIR  = RUN / "_finetune"
PF_PATH = RUN / "_per_fruit" / "per_fruit.json"
PF_PNG  = RUN / "_per_fruit" / "per_fruit.png"
EA_PATH = RUN / "_error_analysis" / "error_analysis.json"
EA_PNG  = RUN / "_error_analysis" / "error_analysis.png"
ENS_DIR = RUN / "_ensemble"

# ── Styles ────────────────────────────────────────────────────
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")   # >= 90%
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")   # 80–89%
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")   # < 80%
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT  = Font(bold=True, size=13)
THIN        = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def f1_fill(pct):
    if pct is None:
        return None
    if pct >= 90: return FILL_GREEN
    if pct >= 80: return FILL_YELLOW
    return FILL_RED


def header(ws, row, cols, widths=None):
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 26


def embed(ws, path, cell, target_w=560):
    """Embed a PNG preserving aspect ratio, anchored at `cell`."""
    if not Path(path).exists():
        ws[cell] = "[plot not found]"
        return
    img = XLImage(str(path))
    ratio = target_w / float(img.width)
    img.width  = target_w
    img.height = int(img.height * ratio)
    ws.add_image(img, cell)


# ── Loaders ───────────────────────────────────────────────────

def load_finetune():
    out = []
    if FT_DIR.exists():
        for f in sorted(FT_DIR.glob("*.json")):
            d = json.load(open(f))
            d["_png"] = FT_DIR / (f.stem + ".png")
            out.append(d)
    # best result first
    out.sort(key=lambda d: d["finetuned"]["f1_macro"], reverse=True)
    return out


def load_json(path):
    return json.load(open(path)) if path.exists() else None


def load_ensemble():
    out = []
    if ENS_DIR.exists():
        for f in sorted(ENS_DIR.glob("*.json")):
            out.append(json.load(open(f)))
    out.sort(key=lambda d: d["ensemble"]["f1_macro"], reverse=True)
    return out


# ── Sheets ────────────────────────────────────────────────────

def write_summary(ws, ft_list, per_fruit, ensemble):
    ws.title = "Summary"
    ws["A1"] = "own_dataset — Domain Adaptation Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    row = 3
    def line(label, value, bold=False):
        nonlocal row
        a = ws.cell(row=row, column=1, value=label)
        a.font = Font(bold=True)
        b = ws.cell(row=row, column=2, value=value)
        if bold:
            b.font = Font(bold=True)
        row += 1

    if ft_list:
        best = ft_list[0]
        bl, fn = best["baseline"], best["finetuned"]
        line("Best base model", best["source_id"])
        line("Photos (k-fold CV)", f"{best['n_photos']}  ({best['folds']} folds)")
        line("Baseline F1", f"{bl['f1_macro']*100:.1f}%")
        line("Fine-tuned F1", f"{fn['f1_macro']*100:.1f}%", bold=True)
        line("Improvement", f"{best['delta_f1_pts']:+.1f} pts  "
                            f"(MCC {bl['mcc']:.3f} → {fn['mcc']:.3f})", bold=True)
        line("Recall (rotten)", f"{bl['recall_rotten']*100:.1f}% → {fn['recall_rotten']*100:.1f}%")
        row += 1

    if per_fruit:
        rows = [m for m in per_fruit["models"]
                if m.get("f1_overall") is not None]
        if rows:
            best_gen = max(rows, key=lambda m: min(
                (m.get("f1_" + fr) or 0) for fr in per_fruit["fruits"]))
            line("Fruits in own_dataset", ", ".join(per_fruit["fruits"]))
            line("Best all-round model", best_gen["id"])
            for fr in per_fruit["fruits"]:
                vals = [m["f1_" + fr] for m in rows if m.get("f1_" + fr) is not None]
                if vals:
                    line(f"Fleet mean F1 — {fr}", f"{sum(vals)/len(vals)*100:.1f}%")
            row += 1

    if ensemble:
        be = ensemble[0]
        line("Best ensemble", f"{be['tag']} / {be['method']}")
        line("Ensemble F1 vs best single",
             f"{be['ensemble']['f1_macro']*100:.1f}%  ({be['delta_f1_pts']:+.1f} pts)")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42


def write_finetune(ws, ft_list):
    ws.title = "Fine-tuning"
    cols = ["Source model", "Config", "Metric", "Baseline", "Fine-tuned", "Δ"]
    header(ws, 1, cols, [18, 26, 16, 12, 12, 10])
    r = 2
    img_anchor_rows = []
    for d in ft_list:
        cfg = d["ft_config"]
        cfg_str = (f"{'freeze' if cfg['unfreeze']==0 else 'unfreeze'}, "
                   f"{cfg['epochs']} ep, lr {cfg['lr_head']}")
        bl, fn = d["baseline"], d["finetuned"]
        metrics = [("F1 macro", "f1_macro", True), ("Accuracy", "acc", True),
                   ("MCC", "mcc", False), ("Recall rotten", "recall_rotten", True)]
        start = r
        for name, key, is_pct in metrics:
            ws.cell(row=r, column=3, value=name)
            if is_pct:
                bv, fv = bl[key]*100, fn[key]*100
                bc = ws.cell(row=r, column=4, value=round(bv, 1)); bc.fill = f1_fill(bv)
                fc = ws.cell(row=r, column=5, value=round(fv, 1)); fc.fill = f1_fill(fv)
                ws.cell(row=r, column=6, value=f"{fv-bv:+.1f}")
            else:
                ws.cell(row=r, column=4, value=round(bl[key], 3))
                ws.cell(row=r, column=5, value=round(fn[key], 3))
                ws.cell(row=r, column=6, value=f"{fn[key]-bl[key]:+.3f}")
            r += 1
        ws.cell(row=start, column=1, value=d["source_id"]).font = Font(bold=True)
        folds = ", ".join(f"{f['f1_ft']*100:.0f}" for f in d["per_fold"])
        ws.cell(row=start, column=2, value=cfg_str)
        ws.cell(row=start+1, column=2, value=f"folds: [{folds}]")
        img_anchor_rows.append((start, d["_png"]))
        r += 1  # blank spacer row

    # embed plots to the right
    for start, png in img_anchor_rows:
        embed(ws, png, f"H{start}", target_w=460)
    for col in "ABCDEF":
        if col != "B":
            ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 8, 12)


def write_per_fruit(ws, pf):
    ws.title = "Per-fruit"
    fruits = pf["fruits"]
    cols = ["ID", "Trained on", "Saw banana"] + [f"F1 {fr}" for fr in fruits] + ["F1 overall"]
    header(ws, 1, cols, [16, 22, 11] + [11]*len(fruits) + [11])
    rows = [m for m in pf["models"] if m.get("f1_overall") is not None]
    rows.sort(key=lambda m: m["f1_overall"], reverse=True)
    r = 2
    for m in rows:
        ws.cell(row=r, column=1, value=m["id"])
        ws.cell(row=r, column=2, value=m["dataset"])
        ws.cell(row=r, column=3, value="yes" if m["saw_banana"] else "no")
        for i, fr in enumerate(fruits):
            v = m.get("f1_" + fr)
            c = ws.cell(row=r, column=4+i, value=round(v*100, 1) if v is not None else "n/a")
            if v is not None:
                c.fill = f1_fill(v*100)
        ov = ws.cell(row=r, column=4+len(fruits), value=round(m["f1_overall"]*100, 1))
        ov.fill = f1_fill(m["f1_overall"]*100)
        r += 1
    embed(ws, PF_PNG, f"{get_column_letter(6+len(fruits))}2", target_w=480)


def write_hardest(ws, ea, top=20):
    ws.title = "Hardest images"
    ws["A1"] = (f"{ea['n_images']} images × {ea['n_models']} models  |  "
                f"accuracy ceiling {ea['accuracy_ceiling']:.1f}%  |  "
                f"images every model fails: {ea['n_unanimous_fail']}")
    ws["A1"].font = Font(bold=True)
    ws.merge_cells("A1:E1")
    header(ws, 3, ["#", "File", "True label", "Models wrong", "Conf when wrong"],
           [5, 22, 12, 16, 16])
    hard = sorted(ea["per_image"], key=lambda s: s["n_wrong"], reverse=True)[:top]
    for i, s in enumerate(hard, 1):
        r = 3 + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=s["file"])
        ws.cell(row=r, column=3, value=s["true_label"])
        c = ws.cell(row=r, column=4, value=f"{s['n_wrong']}/{s['n_models']}")
        c.fill = FILL_RED if s["err_rate"] >= 0.5 else FILL_YELLOW if s["err_rate"] >= 0.1 else None
        ws.cell(row=r, column=5, value=f"{s['conf_wrong']*100:.1f}%")
    embed(ws, EA_PNG, "G3", target_w=620)


def write_ensemble(ws, ens):
    ws.title = "Ensemble"
    header(ws, 1, ["Selection", "Method", "Members", "Best single",
                   "Ensemble F1", "Δ vs best", "MCC", "AUC"],
           [16, 12, 9, 16, 12, 11, 9, 9])
    r = 2
    for d in ens:
        e, bs = d["ensemble"], d["best_single"]
        ws.cell(row=r, column=1, value=d["tag"])
        ws.cell(row=r, column=2, value=d["method"])
        ws.cell(row=r, column=3, value=d["n_members"])
        ws.cell(row=r, column=4, value=f"{bs['id']} ({bs['f1']*100:.1f}%)")
        c = ws.cell(row=r, column=5, value=round(e["f1_macro"]*100, 1)); c.fill = f1_fill(e["f1_macro"]*100)
        ws.cell(row=r, column=6, value=f"{d['delta_f1_pts']:+.1f}")
        ws.cell(row=r, column=7, value=round(e["mcc"], 3))
        ws.cell(row=r, column=8, value=e["auc_roc"])
        r += 1
    best_png = ENS_DIR / "ensemble_top3_mean.png"
    embed(ws, best_png, f"J2", target_w=460)


# ── Main ──────────────────────────────────────────────────────

def main():
    print("\n── Building own_dataset domain-adaptation report ──\n")
    ft   = load_finetune()
    pf   = load_json(PF_PATH)
    ea   = load_json(EA_PATH)
    ens  = load_ensemble()

    found = []
    if ft:  found.append(f"{len(ft)} fine-tuning run(s)")
    if pf:  found.append(f"per-fruit ({len(pf['models'])} models, {len(pf['fruits'])} fruits)")
    if ea:  found.append(f"error analysis ({ea['n_images']} imgs)")
    if ens: found.append(f"{len(ens)} ensemble variant(s)")
    if not found:
        print("No analysis JSON found in run_outputs/_finetune|_per_fruit|"
              "_error_analysis|_ensemble. Run those analyses first.")
        return
    print("  Found: " + " | ".join(found))

    wb = Workbook()
    write_summary(wb.active, ft, pf, ens)
    if ft:  write_finetune(wb.create_sheet("Fine-tuning"), ft)
    if pf:  write_per_fruit(wb.create_sheet("Per-fruit"), pf)
    if ea:  write_hardest(wb.create_sheet("Hardest images"), ea)
    if ens: write_ensemble(wb.create_sheet("Ensemble"), ens)

    wb.save(OUT)
    print(f"\n  Saved: {OUT}  ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
