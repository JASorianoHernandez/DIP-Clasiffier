"""
generate_tracker.py — Generate experiments_tracker.xlsx from existing run_outputs/.

Structure:
  One sheet per backbone + one SUMMARY sheet.
  Rows: dataset x label_mode combinations.
  Columns: C1, C2, C3, C4.
  Cell values: "XX.X% F1" if complete, "pending", "partial", "n/a".

Usage:
    python generate_tracker.py
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────
RUN_DIR  = Path("run_outputs")
OUT_FILE = Path("experiments_tracker.xlsx")

BACKBONES = [
    "resnet18",
    "resnet34",
    "resnet50",
    "mobilenet_v3_small",
    "efficientnet_b0",
    "efficientnet_b2",
]

# Ordered smallest to largest (by image count)
DATASETS = [
    "kaggle_fruits_quality",      #    359 imgs
    "mendeley_fruits",            #  1,655 imgs
    "mendeley_lemon_varieties",   #  1,956 imgs
    "mendeley_fruitvision",       # 10,154 imgs
    "kaggle_fruits_fresh_rotten", # 13,599 imgs
    "kaggle_fresh_stale",         # 27,317 imgs
]

DATASET_ORDER = {ds: i+1 for i, ds in enumerate(DATASETS)}

# datasets where fruit_state doesn't apply
FRUIT_STATE_NA = {
    "mendeley_lemon_varieties",  # single fruit
    "kaggle_fruits_quality",     # flat binary dataset
}

CONDITIONS   = ["frozen", "layer4", "head_frozen", "head_layer4"]
COND_LABELS  = ["C1 frozen", "C2 layer4", "C3 head_frozen", "C4 head_layer4"]
LABEL_MODES  = ["state", "fruit_state"]

# ── Colors ────────────────────────────────────────────────────
COLOR_COMPLETE  = "C6EFCE"   # green
COLOR_PARTIAL   = "FFEB9C"   # yellow
COLOR_PENDING   = "FFFFFF"   # white
COLOR_NA        = "D9D9D9"   # grey
COLOR_HEADER    = "1F4E79"   # dark blue
COLOR_SUBHEADER = "2E75B6"   # medium blue
COLOR_DS_ROW    = "DEEAF1"   # light blue

# ── Load all metrics ──────────────────────────────────────────

def load_runs():
    """
    Return dict keyed by (dataset, condition, backbone, label_mode) → metrics.
    Only counts runs with new metrics format (has best_mcc).
    Old format runs and incomplete runs are ignored (treated as pending).
    Runs with checkpoint but no metrics.json are marked as 'running'.
    """
    runs = {}
    for run_dir in RUN_DIR.iterdir():
        if not run_dir.is_dir() or run_dir.name == "_plots":
            continue
        mpath = run_dir / "metrics.json"
        cpath = run_dir / "checkpoint.pt"

        if not mpath.exists():
            # Has checkpoint but no metrics = currently running
            if cpath.exists():
                runs[(run_dir.name, None, None, None)] = {"status": "running"}
            continue

        with open(mpath) as f:
            m = json.load(f)

        # Skip old format runs (no best_mcc = needs rerun)
        if "best_mcc" not in m:
            continue

        key = (
            m.get("dataset", ""),
            m.get("condition", ""),
            m.get("backbone_name", "resnet18"),
            m.get("label_mode", "state") or "state",
        )
        evals = len(m.get("acc_history", []))
        f1    = m.get("best_f1", 0)
        runs[key] = {
            "status"   : "complete" if evals >= 10 else "partial",
            "best_f1"  : f1,
            "best_acc" : m.get("best_acc", 0),
            "best_mcc" : m.get("best_mcc", None),
            "best_auc" : m.get("best_auc", None),
            "evals"    : evals,
            "time_min" : round(m.get("training_time_seconds", 0) / 60, 1),
        }
    return runs


def is_running(runs, dataset, condition, backbone, label_mode):
    """Check if a run is currently in progress (checkpoint exists, no metrics)."""
    # Match by run folder name pattern
    mode_suffix = "_fs" if label_mode == "fruit_state" else ""
    run_name = f"{dataset}_all_{condition}_{backbone}{mode_suffix}"
    return (run_name, None, None, None) in runs


def cell_value(runs, dataset, condition, backbone, label_mode):
    """Return display string and status for a cell."""
    # n/a cases
    if label_mode == "fruit_state" and dataset in FRUIT_STATE_NA:
        return "n/a", "na"

    key = (dataset, condition, backbone, label_mode)
    if key not in runs:
        # Check if currently running
        if is_running(runs, dataset, condition, backbone, label_mode):
            return "running", "running"
        return "pending", "pending"

    r = runs[key]
    if r["status"] == "partial":
        return "partial", "partial"

    f1  = r["best_f1"]
    acc = r["best_acc"]
    mcc = r["best_mcc"]
    auc = r["best_auc"]
    t   = r["time_min"]

    mcc_str = f"{mcc:.3f}" if mcc is not None else "n/a"
    auc_str = f"{auc:.4f}" if auc is not None else "n/a"

    text = (f"F1:  {f1*100:.1f}%\n"
            f"ACC: {acc*100:.1f}%\n"
            f"MCC: {mcc_str}\n"
            f"AUC: {auc_str}\n"
            f"Time: {t} min")
    return text, "complete"


# ── Style helpers ─────────────────────────────────────────────

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def style_cell(cell, value, status, bold=False):
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()
    if status == "complete":
        cell.fill = make_fill(COLOR_COMPLETE)
    elif status == "partial":
        cell.fill = make_fill(COLOR_PARTIAL)
    elif status == "na":
        cell.fill = make_fill(COLOR_NA)
        cell.font = Font(color="888888", italic=True)
    else:
        cell.fill = make_fill(COLOR_PENDING)
    if bold:
        cell.font = Font(bold=True)


def write_backbone_sheet(ws, backbone, runs):
    """Write one backbone sheet — minimal, no colors."""

    # ── Header row ──
    headers = ["Order", "Dataset", "Label Mode", "C1 frozen", "C2 layer4",
               "C3 head_frozen", "C4 head_layer4"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    # ── Data rows ──
    row = 2
    for ds in DATASETS:
        order = DATASET_ORDER[ds]
        for lm in LABEL_MODES:
            ws.cell(row=row, column=1, value=order)
            ws.cell(row=row, column=2, value=ds)
            ws.cell(row=row, column=3, value=lm)

            for col, cond in enumerate(CONDITIONS, 4):
                val, _ = cell_value(runs, ds, cond, backbone, lm)
                c = ws.cell(row=row, column=col, value=val)
                c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            row += 1

    # ── Column widths ──
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    for col in range(4, 8):
        ws.column_dimensions[get_column_letter(col)].width = 16


def write_summary_sheet(ws, runs):
    """Write summary sheet — best F1 per dataset across all backbones (C4, state)."""

    # Header
    headers = ["Order", "Dataset"] + BACKBONES
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    # Data
    for row, ds in enumerate(DATASETS, 2):
        ws.cell(row=row, column=1, value=DATASET_ORDER[ds])
        ws.cell(row=row, column=2, value=ds)

        for col, bb in enumerate(BACKBONES, 3):
            val, _ = cell_value(runs, ds, "head_layer4", bb, "state")
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 32
    for col in range(3, len(BACKBONES) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 20


# ── Legend sheet ──────────────────────────────────────────────

def write_legend_sheet(ws):
    ws.cell(row=1, column=1, value="Value").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Meaning").font = Font(bold=True)

    items = [
        ("XX.X% F1", "Experiment complete with new metrics format"),
        ("running",  "Experiment currently in progress"),
        ("partial",  "Experiment started but interrupted"),
        ("pending",  "Experiment not yet run"),
        ("n/a",      "Not applicable for this dataset / label mode combination"),
    ]

    for row, (val, desc) in enumerate(items, 2):
        ws.cell(row=row, column=1, value=val)
        ws.cell(row=row, column=2, value=desc)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55


# ── Main ──────────────────────────────────────────────────────

def main():
    runs = load_runs()
    print(f"Loaded {len(runs)} run records from {RUN_DIR}/")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # One sheet per backbone
    for bb in BACKBONES:
        ws = wb.create_sheet(title=bb)
        write_backbone_sheet(ws, bb, runs)
        print(f"  Sheet: {bb}")

    # Summary sheet
    ws_sum = wb.create_sheet(title="SUMMARY")
    write_summary_sheet(ws_sum, runs)
    print(f"  Sheet: SUMMARY")

    # Legend sheet
    ws_leg = wb.create_sheet(title="Legend")
    write_legend_sheet(ws_leg)
    print(f"  Sheet: Legend")

    wb.save(OUT_FILE)
    print(f"\nSaved: {OUT_FILE}")


if __name__ == "__main__":
    main()
