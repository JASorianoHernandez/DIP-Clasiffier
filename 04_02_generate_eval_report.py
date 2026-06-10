"""
04_02_generate_eval_report.py — Generate 04_02_eval_report.xlsx from evaluate.py results.

Sheets:
  {dataset_name}       — one row per model with metrics + embedded plot image
  RANKING_performance  — ranked by F1 on eval dataset (best real-world performance)
  RANKING_robustness   — ranked by Drop (best generalization across domains)
  RANKING_safety       — ranked by Recall (most important for food safety)
  RANKING_accuracy     — ranked by Accuracy (standard metric, informational)
  Legend               — explanation of columns and color coding

Usage:
    python generate_eval_report.py
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── Short ID mappings (same as generate_tracker.py) ──────────
DS_CODE = {
    "kaggle_fruits_quality"     : "KFQ",
    "mendeley_fruits"           : "MFR",
    "mendeley_lemon_varieties"  : "MLM",
    "mendeley_fruitvision"      : "MFV",
    "kaggle_fruits_fresh_rotten": "KFR",
    "kaggle_fresh_stale"        : "KFS",
    "own_dataset"               : "OWN",
}
BB_CODE = {
    "resnet18"          : "R18",
    "mobilenet_v3_small": "MN3",
    "efficientnet_b0"   : "EB0",
    "efficientnet_b2"   : "EB2",
}
COND_CODE = {
    "frozen"     : "C1",
    "layer4"     : "C2",
    "head_frozen": "C3",
    "head_layer4": "C4",
}
MODE_CODE = {
    "state"      : "ST",
    "fruit_state": "FS",
}

def make_id(dataset, backbone, condition, label_mode):
    ds = DS_CODE.get(dataset, dataset[:3].upper())
    bb = BB_CODE.get(backbone, backbone[:3].upper())
    cc = COND_CODE.get(condition, condition)
    lm = MODE_CODE.get(label_mode, label_mode[:2].upper())
    return f"{ds}-{bb}-{cc}-{lm}"

# ── Color fills ───────────────────────────────────────────────
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")  # >= 90%
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")  # 80–89%
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")  # < 80%
FILL_NONE   = PatternFill("none")


def get_fill(value, metric_type="pct"):
    """
    Return fill color based on value and metric type.

    metric_type:
      "pct"       — percentage 0-100, higher is better
      "unit"      — 0 to 1 scale (MCC, AUC), higher is better
      "ms"        — milliseconds, lower is better
      "conf_wrong"— percentage, lower is better (high = overconfident on errors)
    """
    if value is None or value == "n/a":
        return FILL_NONE

    try:
        v = float(value)
    except (TypeError, ValueError):
        return FILL_NONE

    if metric_type == "pct":
        if v >= 90:   return FILL_GREEN
        if v >= 80:   return FILL_YELLOW
        return FILL_RED

    if metric_type == "unit":
        if v >= 0.90: return FILL_GREEN
        if v >= 0.80: return FILL_YELLOW
        return FILL_RED

    if metric_type == "ms":
        if v <= 30:   return FILL_GREEN
        if v <= 60:   return FILL_YELLOW
        return FILL_RED

    if metric_type == "conf_wrong":
        if v <= 60:   return FILL_GREEN
        if v <= 75:   return FILL_YELLOW
        return FILL_RED

    return FILL_NONE

# ── Config ────────────────────────────────────────────────────
RUN_DIR  = Path("run_outputs")
OUT_FILE = Path("04_02_eval_report.xlsx")

# Image size in Excel (pixels)
IMG_WIDTH_PX  = 560
IMG_HEIGHT_PX = 400
ROW_HEIGHT_PT = 300   # row height in points (~400px)

# ── Data loading ──────────────────────────────────────────────

def load_eval_results():
    """
    Scan all run_outputs/{run}/eval/*.json files.
    Returns list of dicts with metrics + paths.
    """
    results = []
    for run_dir in sorted(RUN_DIR.iterdir()):
        if not run_dir.is_dir() or run_dir.name == "_plots":
            continue

        eval_dir = run_dir / "eval"
        if not eval_dir.exists():
            continue

        # Load training metrics for F1_train
        train_metrics_path = run_dir / "metrics.json"
        f1_train = acc_train = prec_train = rec_train = mcc_train = auc_train = None
        backbone_name = condition_name = train_dataset = None
        if train_metrics_path.exists():
            with open(train_metrics_path) as f:
                tm = json.load(f)
            f1_train      = tm.get("best_f1")
            acc_train     = tm.get("best_acc")
            prec_train    = tm.get("best_precision")
            rec_train     = tm.get("best_recall")
            mcc_train     = tm.get("best_mcc")
            auc_train     = tm.get("best_auc")
            backbone_name = tm.get("backbone_name", "resnet18")
            condition_name = tm.get("condition", "")
            train_dataset  = tm.get("dataset", "")

        # Load each eval result
        for eval_path in sorted(eval_dir.glob("*.json")):
            if "preds" in eval_path.name:
                continue
            with open(eval_path) as f:
                em = json.load(f)

            # Find corresponding plot
            plot_name = eval_path.stem + "_plots.png"
            plot_path = eval_dir / plot_name

            # eval metrics
            f1_eval   = round(em.get("f1_macro", 0) * 100, 1)
            acc_eval  = round(em.get("acc", 0) * 100, 1)
            prec_eval = round(em.get("precision_macro", 0) * 100, 1)
            rec_eval  = round(em.get("recall_macro", 0) * 100, 1)
            mcc_eval  = round(em.get("mcc", 0), 3)
            auc_eval  = round(em.get("auc_roc", 0), 4) if em.get("auc_roc") else None

            def pct(v):   return round(v * 100, 1) if v is not None else None
            def drop(t,e): return round(t - e, 1) if t is not None else None

            # ID must encode the dataset the model was TRAINED on, not the
            # dataset it is being evaluated on (own_dataset). The eval dataset
            # is shown separately in its own column.
            run_id = make_id(
                train_dataset or em.get("eval_dataset", ""),
                backbone_name or "resnet18",
                condition_name or "",
                em.get("label_mode", "state"),
            )

            results.append({
                "id"           : run_id,
                "model_run"    : run_dir.name,
                "eval_dataset" : em.get("eval_dataset", ""),
                "label_mode"   : em.get("label_mode", ""),
                # train metrics
                "f1_train"     : pct(f1_train),
                "acc_train"    : pct(acc_train),
                "prec_train"   : pct(prec_train),
                "rec_train"    : pct(rec_train),
                "mcc_train"    : round(mcc_train, 3) if mcc_train else None,
                "auc_train"    : round(auc_train, 4) if auc_train else None,
                # eval metrics
                "acc"          : acc_eval,
                "f1"           : f1_eval,
                "precision"    : prec_eval,
                "recall"       : rec_eval,
                "mcc"          : mcc_eval,
                "auc"          : auc_eval,
                "ms_per_img"   : em.get("inference_ms_per_img"),
                "conf_correct" : round(em.get("conf_avg_correct", 0) * 100, 1) if em.get("conf_avg_correct") else None,
                "conf_wrong"   : round(em.get("conf_avg_wrong", 0) * 100, 1) if em.get("conf_avg_wrong") else None,
                "plot_path"    : plot_path if plot_path.exists() else None,
                # drops (train - eval, lower is better)
                "drop"         : drop(pct(f1_train),   f1_eval),
                "drop_acc"     : drop(pct(acc_train),  acc_eval),
                "drop_prec"    : drop(pct(prec_train), prec_eval),
                "drop_recall"  : drop(pct(rec_train),  rec_eval),
                "drop_mcc"     : round(mcc_train - mcc_eval, 3) if mcc_train else None,
                "drop_auc"     : round(auc_train - auc_eval, 4) if (auc_train and auc_eval) else None,
            })

    return results


# ── Sheet writers ─────────────────────────────────────────────

METRIC_HEADERS = [
    "ID", "Model", "Label Mode",
    "F1 Train (%)", "F1 Eval (%)", "F1 Drop (%)",
    "Acc Train (%)", "Acc Eval (%)", "Acc Drop (%)",
    "Prec Train (%)", "Prec Eval (%)", "Prec Drop (%)",
    "Rec Train (%)", "Rec Eval (%)", "Rec Drop (%)",
    "MCC Train", "MCC Eval", "MCC Drop",
    "AUC Train", "AUC Eval", "AUC Drop",
    "ms/img", "Conf Correct (%)", "Conf Wrong (%)",
]

METRIC_WIDTHS = [14, 42, 14,
                 13, 12, 10,
                 13, 12, 10,
                 13, 12, 10,
                 12, 12, 10,
                 10, 10, 10,
                 10, 10, 10,
                 10, 16, 14]
IMG_START_COL = len(METRIC_HEADERS) + 2   # one blank column after metrics


def auto_fit_columns(ws, min_width=8, max_width=60, skip_cols=None):
    """Set column widths based on longest content. skip_cols: set of col letters to skip (e.g. image cols)."""
    skip_cols = skip_cols or set()
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter in skip_cols:
            continue
        max_len = 0
        for cell in col:
            if cell.value is not None:
                lines = str(cell.value).split("\n")
                max_len = max(max_len, max(len(line) for line in lines))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def write_header(ws, headers, widths):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 18


def write_eval_sheet(ws, rows, ds_name, mode):
    """Write one eval dataset sheet with metrics + images."""
    ws.title = f"{ds_name} ({mode})"

    write_header(ws, METRIC_HEADERS, METRIC_WIDTHS)

    # Image column width
    img_col_letter = get_column_letter(IMG_START_COL)
    ws.column_dimensions[img_col_letter].width = IMG_WIDTH_PX / 7  # approx pts

    for row_idx, r in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT

        def v(key):   return r[key] if r.get(key) is not None else "n/a"

        values = [
            r["id"], r["model_run"], r["label_mode"],
            v("f1_train"),   v("f1"),        v("drop"),
            v("acc_train"),  v("acc"),        v("drop_acc"),
            v("prec_train"), v("precision"),  v("drop_prec"),
            v("rec_train"),  v("recall"),     v("drop_recall"),
            v("mcc_train"),  v("mcc"),        v("drop_mcc"),
            v("auc_train"),  v("auc"),        v("drop_auc"),
            v("ms_per_img"), v("conf_correct"), v("conf_wrong"),
        ]

        # metric type per column (None = no color)
        metric_types = [
            None, None, None,
            "pct", "pct", "drop",
            "pct", "pct", "drop",
            "pct", "pct", "drop",
            "pct", "pct", "drop",
            "unit", "unit", "drop",
            "unit", "unit", "drop",
            "ms", "pct", "conf_wrong",
        ]

        for col, (val, mtype) in enumerate(zip(values, metric_types), 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            if mtype:
                c.fill = get_fill(val, mtype)

        # Embed plot image
        if r["plot_path"]:
            try:
                img = XLImage(str(r["plot_path"]))
                img.width  = IMG_WIDTH_PX
                img.height = IMG_HEIGHT_PX
                cell_ref = f"{img_col_letter}{row_idx}"
                ws.add_image(img, cell_ref)
            except Exception as e:
                ws.cell(row=row_idx, column=IMG_START_COL,
                        value=f"[image error: {e}]")

    # auto-fit all columns except the image column
    img_col_letter = get_column_letter(IMG_START_COL)
    auto_fit_columns(ws, skip_cols={img_col_letter})


def _write_ranking(ws, title, all_results, sort_key, reverse, headers, widths, rank_types, get_row_values):
    """Generic ranking sheet writer."""
    ws.title = title
    write_header(ws, ["#"] + headers, [5] + widths)
    ws.column_dimensions["A"].width = 5

    ranked = sorted(all_results,
                    key=lambda r: sort_key(r),
                    reverse=reverse)

    for rank, r in enumerate(ranked, start=1):
        row_idx = rank + 1
        values = get_row_values(r)

        # Rank number
        c0 = ws.cell(row=row_idx, column=1, value=rank)
        c0.alignment = Alignment(horizontal="center", vertical="center")
        c0.font = Font(bold=True)

        for col, (val, mtype) in enumerate(zip(values, rank_types), 2):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if mtype == "drop":
                try:
                    v = float(val)
                    if v <= 5:    c.fill = FILL_GREEN
                    elif v <= 15: c.fill = FILL_YELLOW
                    else:         c.fill = FILL_RED
                except (TypeError, ValueError):
                    pass
            elif mtype:
                c.fill = get_fill(val, mtype)
        ws.row_dimensions[row_idx].height = 18

    auto_fit_columns(ws)


def write_ranking_performance(ws, all_results):
    """RANKING by F1_eval. Primary metric (F1) shows Train/Eval/Drop. Rest: Eval only."""
    headers = [
        "ID", "Model", "Eval Dataset", "Label Mode",
        "F1 Train (%)", "F1 Eval (%)", "F1 Drop (%)",
        "Acc Eval (%)", "Prec Eval (%)", "Rec Eval (%)",
        "MCC", "AUC", "ms/img",
        "Conf Correct (%)", "Conf Wrong (%)",
    ]
    widths     = [14,42,18,14, 13,12,10, 12,13,12, 10,10,10, 16,14]
    rank_types = [None,None,None,None, "pct","pct","drop", "pct","pct","pct", "unit","unit","ms", "pct","conf_wrong"]

    def v(r,k): return r[k] if r.get(k) is not None else "n/a"
    def sort_key(r): return r["f1"] if r["f1"] else 0
    def get_row(r): return [
        r["id"], r["model_run"], r["eval_dataset"], r["label_mode"],
        v(r,"f1_train"), v(r,"f1"), v(r,"drop"),
        v(r,"acc"), v(r,"precision"), v(r,"recall"),
        v(r,"mcc"), v(r,"auc"), v(r,"ms_per_img"),
        v(r,"conf_correct"), v(r,"conf_wrong"),
    ]
    _write_ranking(ws, "RANKING_performance", all_results,
                   sort_key, True, headers, widths, rank_types, get_row)


def write_ranking_robustness(ws, all_results):
    """RANKING by F1 Drop. Primary metric (Drop) shows Train/Eval/Drop. Rest: Eval only."""
    headers = [
        "ID", "Model", "Eval Dataset", "Label Mode",
        "F1 Train (%)", "F1 Eval (%)", "F1 Drop (%)",
        "Acc Eval (%)", "Prec Eval (%)", "Rec Eval (%)",
        "MCC", "AUC", "ms/img",
        "Conf Correct (%)", "Conf Wrong (%)",
    ]
    widths     = [14,42,18,14, 13,12,10, 12,13,12, 10,10,10, 16,14]
    rank_types = [None,None,None,None, "pct","pct","drop", "pct","pct","pct", "unit","unit","ms", "pct","conf_wrong"]

    def v(r,k): return r[k] if r.get(k) is not None else "n/a"
    def sort_key(r): return r["drop"] if r["drop"] is not None else 999
    def get_row(r): return [
        r["id"], r["model_run"], r["eval_dataset"], r["label_mode"],
        v(r,"f1_train"), v(r,"f1"), v(r,"drop"),
        v(r,"acc"), v(r,"precision"), v(r,"recall"),
        v(r,"mcc"), v(r,"auc"), v(r,"ms_per_img"),
        v(r,"conf_correct"), v(r,"conf_wrong"),
    ]
    _write_ranking(ws, "RANKING_robustness", all_results,
                   sort_key, False, headers, widths, rank_types, get_row)


def write_ranking_safety(ws, all_results):
    """RANKING by Recall. Primary metric (Recall) shows Train/Eval/Drop. Rest: Eval only."""
    headers = [
        "ID", "Model", "Eval Dataset", "Label Mode",
        "Rec Train (%)", "Rec Eval (%)", "Rec Drop (%)",
        "F1 Eval (%)", "Acc Eval (%)", "Prec Eval (%)",
        "MCC", "AUC", "ms/img",
        "Conf Wrong (%)", "Conf Correct (%)",
    ]
    widths     = [14,42,18,14, 12,12,10, 12,12,13, 10,10,10, 16,16]
    rank_types = [None,None,None,None, "pct","pct","drop", "pct","pct","pct", "unit","unit","ms", "conf_wrong","pct"]

    def v(r,k): return r[k] if r.get(k) is not None else "n/a"
    def sort_key(r): return r["recall"] if r["recall"] else 0
    def get_row(r): return [
        r["id"], r["model_run"], r["eval_dataset"], r["label_mode"],
        v(r,"rec_train"), v(r,"recall"), v(r,"drop_recall"),
        v(r,"f1"), v(r,"acc"), v(r,"precision"),
        v(r,"mcc"), v(r,"auc"), v(r,"ms_per_img"),
        v(r,"conf_wrong"), v(r,"conf_correct"),
    ]
    _write_ranking(ws, "RANKING_safety", all_results,
                   sort_key, True, headers, widths, rank_types, get_row)


def write_ranking_accuracy(ws, all_results):
    """RANKING by Accuracy. Primary metric (Acc) shows Train/Eval/Drop. Rest: Eval only."""
    headers = [
        "ID", "Model", "Eval Dataset", "Label Mode",
        "Acc Train (%)", "Acc Eval (%)", "Acc Drop (%)",
        "F1 Eval (%)", "Prec Eval (%)", "Rec Eval (%)",
        "MCC", "AUC", "ms/img",
        "Conf Correct (%)", "Conf Wrong (%)",
    ]
    widths     = [14,42,18,14, 13,12,10, 12,13,12, 10,10,10, 16,14]
    rank_types = [None,None,None,None, "pct","pct","drop", "pct","pct","pct", "unit","unit","ms", "pct","conf_wrong"]

    def v(r,k): return r[k] if r.get(k) is not None else "n/a"
    def sort_key(r): return r["acc"] if r["acc"] else 0
    def get_row(r): return [
        r["id"], r["model_run"], r["eval_dataset"], r["label_mode"],
        v(r,"acc_train"), v(r,"acc"), v(r,"drop_acc"),
        v(r,"f1"), v(r,"precision"), v(r,"recall"),
        v(r,"mcc"), v(r,"auc"), v(r,"ms_per_img"),
        v(r,"conf_correct"), v(r,"conf_wrong"),
    ]
    _write_ranking(ws, "RANKING_accuracy", all_results,
                   sort_key, True, headers, widths, rank_types, get_row)


def write_legend_sheet(ws):
    ws.title = "Legend"
    row = 1

    def section(title):
        nonlocal row
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

    def entry(col1, col2=""):
        nonlocal row
        ws.cell(row=row, column=1, value=col1)
        ws.cell(row=row, column=2, value=col2)
        row += 1

    def blank():
        nonlocal row
        row += 1

    # ── ID format ──
    section("ID FORMAT:  {DATASET}-{BACKBONE}-{CONDITION}-{LABELMODE}")
    entry("Example", "KFQ-R18-C4-ST  =  kaggle_fruits_quality, ResNet18, head_layer4, state")
    entry("Purpose",  "Short unique identifier — search in 04_01_experiments_tracker.xlsx to cross-reference")

    blank()

    section("DATASET CODES")
    for ds, code in DS_CODE.items():
        entry(code, ds)

    blank()

    section("BACKBONE CODES")
    for bb, code in BB_CODE.items():
        entry(code, bb)

    blank()

    section("CONDITION CODES")
    entry("C1", "frozen      — backbone completely frozen, linear head only")
    entry("C2", "layer4      — last backbone block unfrozen, linear head only")
    entry("C3", "head_frozen — backbone frozen, projection head (512→256→128) + linear")
    entry("C4", "head_layer4 — layer4 unfrozen + projection head + linear (best results)")

    blank()

    section("LABEL MODE CODES")
    entry("ST", "state       — classify by freshness state only (fresh / rotten / formalin)")
    entry("FS", "fruit_state — classify by fruit + state (apple_fresh / apple_rotten / ...)")

    blank()

    section("METRIC COLUMNS")
    entry("F1 Train (%)",     "Best macro F1 on the model's own training dataset")
    entry("F1 Eval (%)",      "Macro F1 on the cross-dataset evaluation")
    entry("F1 Drop (%)",      "F1 Train - F1 Eval — lower = better generalization")
    entry("Acc Train/Eval/Drop", "Same as F1 but for accuracy")
    entry("Prec Train/Eval/Drop","Same as F1 but for macro precision")
    entry("Rec Train/Eval/Drop", "Same as F1 but for macro recall — critical for food safety")
    entry("MCC",               "Matthews Correlation Coefficient (-1 worst / 0 random / +1 perfect)")
    entry("AUC",               "Area Under ROC Curve (0.5 random / 1.0 perfect)")
    entry("ms/img",            "Inference time per image in ms (includes DataLoader + GPU overhead)")
    entry("Conf Correct (%)",  "Average model confidence on correct predictions. "
                               "Computed as mean(max(softmax(logits))) for images where pred == true label. "
                               "Higher is better — the model should be certain when it is right.")
    entry("Conf Wrong (%)",    "Average model confidence on wrong predictions. "
                               "Computed as mean(max(softmax(logits))) for images where pred != true label. "
                               "Lower is better. High Conf Wrong (>75%) means the model is overconfident "
                               "on its errors — dangerous in food safety since a wrong prediction with "
                               "95% confidence raises no alert.")

    blank()

    section("HOW CONFIDENCE IS CALCULATED (Confidence Distribution Chart)")
    entry("Step 1 — Logits",
          "The model outputs raw scores (logits) for each class. "
          "Example: [2.3, -0.8] for [fresh, rotten].")
    entry("Step 2 — Softmax",
          "Logits are converted to probabilities: P(c) = exp(z_c) / sum(exp(z)). "
          "Example: [2.3, -0.8] -> [95.7%, 4.3%]. Always sums to 100%.")
    entry("Step 3 — Confidence",
          "Confidence = max(softmax) = 95.7%. The model predicts the class with "
          "the highest probability.")
    entry("Step 4 — Histogram",
          "For each image, confidence is recorded and labeled correct (green) or "
          "wrong (red). The chart shows how many images fell in each confidence "
          "range. Ideal: green near 100%, red near 50-60%. "
          "Warning: red bars near 100% mean the model is wrong with high certainty.")

    blank()

    section("RANKING SHEETS")
    entry("RANKING_performance", "Ordered by F1 Eval descending — best real-world performance")
    entry("RANKING_robustness",  "Ordered by F1 Drop ascending — best generalization across domains")
    entry("RANKING_safety",      "Ordered by Recall Eval descending — most critical for food safety")
    entry("RANKING_accuracy",    "Ordered by Accuracy Eval descending — standard metric (informational)")
    entry("Note", "Primary ranking metric shows Train / Eval / Drop. All other metrics show Eval only.")

    blank()

    section("COLOR CODING")
    entry("% metrics (F1, Acc, Prec, Recall, Conf Correct)", "Green >= 90%  /  Yellow 80-89%  /  Red < 80%")
    entry("Unit metrics (MCC, AUC)",                          "Green >= 0.90  /  Yellow 0.80-0.89  /  Red < 0.80")
    entry("ms/img — lower is better",                         "Green <= 30ms  /  Yellow 30-60ms  /  Red > 60ms")
    entry("Conf Wrong — lower is better",                     "Green <= 60%  /  Yellow 60-75%  /  Red > 75%")
    entry("Drop — lower is better",                           "Green <= 5%  /  Yellow 5-15%  /  Red > 15%")

    auto_fit_columns(ws)


# ── Main ──────────────────────────────────────────────────────

def main():
    results = load_eval_results()

    if not results:
        print("No eval results found. Run evaluate.py first.")
        return

    print(f"Loaded {len(results)} eval records.")

    # Group by eval_dataset + label_mode
    groups = {}
    for r in results:
        key = (r["eval_dataset"], r["label_mode"])
        groups.setdefault(key, []).append(r)

    # Sort each group by F1 descending
    for key in groups:
        groups[key].sort(key=lambda r: r["f1"], reverse=True)

    wb = Workbook()
    wb.remove(wb.active)

    # One sheet per eval dataset+mode
    for (ds_name, mode), rows in sorted(groups.items()):
        ws = wb.create_sheet()
        write_eval_sheet(ws, rows, ds_name, mode)
        print(f"  Sheet: {ds_name} ({mode}) — {len(rows)} models")

    # Ranking sheets
    write_ranking_performance(wb.create_sheet(), results)
    print(f"  Sheet: RANKING_performance")
    write_ranking_robustness(wb.create_sheet(), results)
    print(f"  Sheet: RANKING_robustness")
    write_ranking_safety(wb.create_sheet(), results)
    print(f"  Sheet: RANKING_safety")
    write_ranking_accuracy(wb.create_sheet(), results)
    print(f"  Sheet: RANKING_accuracy")

    # Legend sheet
    ws_leg = wb.create_sheet(title="Legend")
    write_legend_sheet(ws_leg)
    print(f"  Sheet: Legend")

    wb.save(OUT_FILE)
    print(f"\nSaved: {OUT_FILE}")


if __name__ == "__main__":
    main()
