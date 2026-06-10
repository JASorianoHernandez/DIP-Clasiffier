"""
04_04_generate_benchmark_report.py — Benchmark comparison vs published papers.

Compares our best individual models and ensemble against published papers that
used the same public datasets (KFR: apple/banana/orange × fresh/rotten, 6 classes;
MFR: peach/pomegranate/strawberry × fresh/rotten, 6 classes).

Inputs (all auto-discovered):
  run_outputs/<ds>_*_fs/eval/<ds>_fruit_state.json   individual model eval metrics
  run_outputs/<ds>_*_fs/metrics.json                  backbone / condition per model
  run_outputs/_ensemble/ensemble_KFR-FS_all_*.json    KFR ensemble results
  run_outputs/_ensemble/ensemble_MFR-FS_all_*.json    MFR ensemble results
  run_outputs/_ensemble/ensemble_KFR-FS_all_mean.png  KFR ensemble bar chart
  run_outputs/_ensemble/ensemble_MFR-FS_all_mean.png  MFR ensemble bar chart
  [Paper results are hardcoded from published PDFs]

Output: benchmark_report.xlsx
  Sheets: Summary | KFR vs Papers | MFR vs Papers

Usage:
    python 04_04_generate_benchmark_report.py
"""

import sys
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

RUN     = Path("run_outputs")
ENS_DIR = RUN / "_ensemble"
OUT     = Path("04_04_benchmark_report.xlsx")

# ── Dataset config ─────────────────────────────────────────────
DS_CONFIG = {
    "KFR": {
        "name"   : "kaggle_fruits_fresh_rotten",
        "label"  : "KFR — apple / banana / orange (6 classes)",
        "classes": 6,
        "images" : 13_599,
        "split"  : "80 / 20",
    },
    "MFR": {
        "name"   : "mendeley_fruits",
        "label"  : "MFR — peach / pomegranate / strawberry (6 classes)",
        "classes": 6,
        "images" : 1_655,
        "split"  : "80 / 20",
    },
}

# ── Published paper results (hardcoded from PDFs) ──────────────
# All papers report Accuracy as primary metric.
PAPERS = {
    "KFR": [
        {
            "ref"   : "Palakodati et al., 2020",
            "venue" : "Rev. d'Intell. Artif., Vol.34 No.5",
            "model" : "Proposed CNN",
            "size"  : "5,989 (subset)",
            "split" : "60 / 10 / 30",
            "acc"   : 0.9782,
            "f1"    : None,
            "note"  : "Partial KFR; test-set accuracy",
        },
        {
            "ref"   : "Chakraborty et al., 2021",
            "venue" : "IEEE ICOEI",
            "model" : "MobileNetV2",
            "size"  : "13,599 (full)",
            "split" : "80 / 20",
            "acc"   : 0.9961,
            "f1"    : None,
            "note"  : "Full KFR; validation accuracy",
        },
    ],
    "MFR": [
        {
            "ref"   : "Sharma & Kumar, 2025",
            "venue" : "IEEE ICPCT",
            "model" : "ResNet50 (fine-tuned)",
            "size"  : "1,655 (full)",
            "split" : "70 / 15 / 15",
            "acc"   : 0.95,
            "f1"    : 0.95,
            "note"  : "MFR dataset; test-set accuracy",
        },
    ],
}

# ── Code mappings ──────────────────────────────────────────────
BB_CODE = {
    "resnet18"          : "R18",
    "mobilenet_v3_small": "MN3",
    "efficientnet_b0"   : "EB0",
    "efficientnet_b2"   : "EB2",
}
COND_CODE = {
    "frozen"     : "C1",
    "layer4"     : "C2",
    "head_frozen": "C3",
    "head_layer4": "C4",
}

# ── Styles ─────────────────────────────────────────────────────
FILL_GREEN   = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW  = PatternFill("solid", fgColor="FFEB9C")
FILL_RED     = PatternFill("solid", fgColor="FFC7CE")
FILL_GOLD    = PatternFill("solid", fgColor="FFD966")
FILL_BLUE    = PatternFill("solid", fgColor="BDD7EE")
FILL_PAPER   = PatternFill("solid", fgColor="F2F2F2")
FILL_SECTION = PatternFill("solid", fgColor="D9E1F2")
FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
FONT_HEADER  = Font(bold=True, color="FFFFFF")
FONT_TITLE   = Font(bold=True, size=13)
FONT_BOLD    = Font(bold=True)
FONT_NOTE    = Font(italic=True, size=9, color="595959")
THIN         = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def acc_fill(v):
    if v is None: return None
    if v >= 0.99: return FILL_GREEN
    if v >= 0.95: return FILL_YELLOW
    return FILL_RED


def hdr(ws, row, cols, widths=None):
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 26


def section_row(ws, row, text, n_cols):
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_BOLD
    c.fill = FILL_SECTION
    if n_cols > 1:
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")


def embed_png(ws, path, cell, target_w=520):
    if not Path(path).exists():
        ws[cell] = "[plot not found]"
        return
    img = XLImage(str(path))
    ratio    = target_w / float(img.width)
    img.width  = target_w
    img.height = int(img.height * ratio)
    ws.add_image(img, cell)


# ── Data loaders ───────────────────────────────────────────────

def load_models(ds_code):
    """Load individual model eval results for one dataset (fruit_state)."""
    ds_name = DS_CONFIG[ds_code]["name"]
    models  = []
    for run_dir in sorted(RUN.iterdir()):
        if not run_dir.is_dir() or run_dir.name.startswith("_"):
            continue
        if not (run_dir.name.startswith(ds_name) and run_dir.name.endswith("_fs")):
            continue
        eval_path = run_dir / "eval" / f"{ds_name}_fruit_state.json"
        mpath     = run_dir / "metrics.json"
        if not (eval_path.exists() and mpath.exists()):
            continue
        ev = json.load(open(eval_path))
        tr = json.load(open(mpath))
        bb   = tr.get("backbone_name", "")
        cond = tr.get("condition", "")
        models.append({
            "id"  : (f"{ds_code}-{BB_CODE.get(bb, bb[:3].upper())}"
                     f"-{COND_CODE.get(cond, cond)}-FS"),
            "bb"  : BB_CODE.get(bb, bb),
            "cond": COND_CODE.get(cond, cond),
            "acc" : ev.get("acc", 0.0),
            "f1"  : ev.get("f1_macro", 0.0),
            "prec": ev.get("precision_macro", 0.0),
            "rec" : ev.get("recall_macro", 0.0),
            "mcc" : ev.get("mcc", 0.0),
            "auc" : ev.get("auc_roc"),
        })
    models.sort(key=lambda m: m["acc"], reverse=True)
    return models


def load_ensembles(ds_code):
    """Load ensemble results for one dataset (all methods, all-models selection)."""
    ensembles = []
    for f in sorted(ENS_DIR.glob(f"ensemble_{ds_code}-FS_all_*.json")):
        d = json.load(open(f))
        ensembles.append({
            "method"         : d["method"],
            "n"              : d["n_members"],
            "acc"            : d["ensemble"]["acc"],
            "f1"             : d["ensemble"]["f1_macro"],
            "mcc"            : d["ensemble"]["mcc"],
            "delta"          : d["delta_f1_pts"],
            "best_single_id" : d["best_single"]["id"],
            "best_single_f1" : d["best_single"]["f1"],
            "png"            : ENS_DIR / f"ensemble_{ds_code}-FS_all_{d['method']}.png",
        })
    ensembles.sort(key=lambda e: e["acc"], reverse=True)
    return ensembles


# ── Sheet: Summary ─────────────────────────────────────────────

def write_summary(ws, all_results):
    ws.title = "Summary"
    ws["A1"] = "Benchmark Summary — Our Models vs Published Papers"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:G1")
    ws["A2"] = ("Primary metric: Accuracy (matches papers). F1 macro shown alongside. "
                "Gold = best result per dataset.")
    ws["A2"].font = FONT_NOTE
    ws.merge_cells("A2:G2")

    COLS   = ["Source", "Model", "Dataset (size)", "Split", "Accuracy", "F1 macro", "Notes"]
    WIDTHS = [18, 30, 20, 12, 12, 12, 32]

    row = 4
    for ds_code, (models, ensembles, papers) in all_results.items():
        cfg      = DS_CONFIG[ds_code]
        best_pub = max(p["acc"] for p in papers)

        # ── Section header ──
        section_row(ws, row, f"  {cfg['label']}  ·  {cfg['images']:,} images", len(COLS))
        row += 1
        hdr(ws, row, COLS, WIDTHS)
        row += 1

        # Papers
        for p in papers:
            ws.cell(row=row, column=1, value="Paper").fill = FILL_PAPER
            ws.cell(row=row, column=2, value=f"{p['model']}  ({p['ref']})").fill = FILL_PAPER
            ws.cell(row=row, column=3, value=f"{p['size']}").fill = FILL_PAPER
            ws.cell(row=row, column=4, value=p["split"]).fill = FILL_PAPER
            c = ws.cell(row=row, column=5, value=f"{p['acc']*100:.2f}%")
            c.fill = FILL_PAPER
            ws.cell(row=row, column=6,
                    value=f"{p['f1']*100:.2f}%" if p["f1"] else "—").fill = FILL_PAPER
            ws.cell(row=row, column=7, value=p.get("note", "")).fill = FILL_PAPER
            row += 1

        # Best single model
        if models:
            best = models[0]
            ws.cell(row=row, column=1, value="Ours — best single").font = FONT_BOLD
            ws.cell(row=row, column=2, value=best["id"]).font = FONT_BOLD
            ws.cell(row=row, column=3, value=f"{cfg['images']:,}  (20 % test)")
            ws.cell(row=row, column=4, value=cfg["split"])
            is_best = best["acc"] > best_pub
            ca = ws.cell(row=row, column=5, value=f"{best['acc']*100:.2f}%")
            ca.fill = FILL_GOLD if is_best else FILL_GREEN
            ca.font = FONT_BOLD
            cf = ws.cell(row=row, column=6, value=f"{best['f1']*100:.2f}%")
            cf.fill = FILL_GOLD if is_best else FILL_GREEN
            delta = (best["acc"] - best_pub) * 100
            ws.cell(row=row, column=7,
                    value=f"Transfer learning · C3 frozen backbone + projection head · Δ {delta:+.2f} pts vs best paper")
            row += 1

        # Best ensemble
        if ensembles:
            be = ensembles[0]
            ws.cell(row=row, column=1, value=f"Ours — ensemble ({be['n']} models)").font = FONT_BOLD
            ws.cell(row=row, column=2,
                    value=f"16 models × 4 arch × 4 cond  ({be['method']})").font = FONT_BOLD
            ws.cell(row=row, column=3, value=f"{cfg['images']:,}  (20 % test)")
            ws.cell(row=row, column=4, value=cfg["split"])
            is_best = be["acc"] > best_pub
            ca = ws.cell(row=row, column=5, value=f"{be['acc']*100:.2f}%")
            ca.fill = FILL_GOLD if is_best else FILL_GREEN
            ca.font = FONT_BOLD
            cf = ws.cell(row=row, column=6, value=f"{be['f1']*100:.2f}%")
            cf.fill = FILL_GOLD if is_best else FILL_GREEN
            delta = (be["acc"] - best_pub) * 100
            ws.cell(row=row, column=7,
                    value=f"Δ {delta:+.2f} pts vs best paper · {be['delta']:+.2f} pts vs best single")
            row += 1

        row += 1  # spacer between datasets


# ── Sheet: per-dataset detail ──────────────────────────────────

def write_ds_sheet(ws, ds_code, models, ensembles, papers):
    cfg   = DS_CONFIG[ds_code]
    N_COL = 8

    ws["A1"] = f"{cfg['label']}  —  Benchmark vs Papers"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(f"A1:{get_column_letter(N_COL)}1")

    row = 3

    # ── Published papers ──
    section_row(ws, row, "  Published Papers", N_COL); row += 1
    hdr(ws, row,
        ["Reference", "Venue", "Model", "Dataset size", "Split",
         "Accuracy", "F1 macro", "Notes"],
        [28, 22, 20, 14, 12, 12, 12, 30])
    row += 1
    best_pub = max(p["acc"] for p in papers)
    for p in papers:
        ws.cell(row=row, column=1, value=p["ref"])
        ws.cell(row=row, column=2, value=p["venue"])
        ws.cell(row=row, column=3, value=p["model"])
        ws.cell(row=row, column=4, value=p["size"])
        ws.cell(row=row, column=5, value=p["split"])
        c = ws.cell(row=row, column=6, value=f"{p['acc']*100:.2f}%")
        c.fill = acc_fill(p["acc"])
        ws.cell(row=row, column=7,
                value=f"{p['f1']*100:.2f}%" if p["f1"] else "—")
        ws.cell(row=row, column=8, value=p.get("note", ""))
        row += 1
    row += 1

    # ── Ensemble results ──
    if ensembles:
        section_row(ws, row, "  Our Ensemble Results  (4 architectures × 4 conditions = 16 models)", N_COL)
        row += 1
        hdr(ws, row,
            ["Method", "Members", "Accuracy", "F1 macro", "MCC",
             "Δ vs best single", "Best single model", "Δ vs best paper"],
            [14, 10, 12, 12, 10, 18, 24, 18])
        row += 1
        for e in ensembles:
            ws.cell(row=row, column=1, value=e["method"])
            ws.cell(row=row, column=2, value=e["n"])
            ca = ws.cell(row=row, column=3, value=f"{e['acc']*100:.2f}%")
            ca.fill = FILL_GOLD if e["acc"] > best_pub else FILL_GREEN
            ca.font = FONT_BOLD
            cf = ws.cell(row=row, column=4, value=f"{e['f1']*100:.2f}%")
            cf.fill = FILL_GOLD if e["acc"] > best_pub else FILL_GREEN
            ws.cell(row=row, column=5, value=round(e["mcc"], 4))
            ws.cell(row=row, column=6, value=f"{e['delta']:+.2f} pts")
            ws.cell(row=row, column=7, value=e["best_single_id"])
            delta_pub = (e["acc"] - best_pub) * 100
            dp = ws.cell(row=row, column=8, value=f"{delta_pub:+.2f} pts")
            dp.font = Font(bold=True,
                           color="375623" if delta_pub >= 0 else "9C0006")
            row += 1

        # Embed ensemble bar-chart plot
        best_png = ensembles[0]["png"]
        embed_png(ws, best_png, f"J3", target_w=520)
        row += 1

    # ── All individual models ranked ──
    section_row(ws, row, f"  All Individual Models — ranked by Accuracy  ({len(models)} models)", N_COL)
    row += 1
    hdr(ws, row,
        ["Rank", "Model ID", "Backbone", "Condition", "Accuracy",
         "F1 macro", "Precision", "Recall"],
        [7, 24, 10, 11, 12, 12, 12, 12])
    row += 1
    for rank, m in enumerate(models, 1):
        ws.cell(row=row, column=1, value=rank)
        id_c = ws.cell(row=row, column=2, value=m["id"])
        if rank == 1:
            id_c.font = FONT_BOLD
        ws.cell(row=row, column=3, value=m["bb"])
        ws.cell(row=row, column=4, value=m["cond"])
        ca = ws.cell(row=row, column=5, value=f"{m['acc']*100:.2f}%")
        ca.fill = acc_fill(m["acc"])
        if rank == 1:
            ca.font = FONT_BOLD
        ws.cell(row=row, column=6, value=f"{m['f1']*100:.2f}%").fill = acc_fill(m["f1"])
        ws.cell(row=row, column=7, value=f"{m['prec']*100:.2f}%")
        ws.cell(row=row, column=8, value=f"{m['rec']*100:.2f}%")
        row += 1


# ── Main ───────────────────────────────────────────────────────

def main():
    print("\n── Building benchmark report ──\n")

    all_results = {}
    for ds_code in ["KFR", "MFR"]:
        models    = load_models(ds_code)
        ensembles = load_ensembles(ds_code)
        papers    = PAPERS[ds_code]

        if not models:
            print(f"  [warn] No {ds_code} fruit_state eval results found — skipping.")
            continue

        best      = models[0]
        best_pub  = max(p["acc"] for p in papers)
        print(f"  {ds_code}: {len(models)} models  |  {len(ensembles)} ensemble variants")
        print(f"    Best single  : {best['id']}  acc={best['acc']*100:.2f}%  "
              f"F1={best['f1']*100:.2f}%  Δ vs best paper: {(best['acc']-best_pub)*100:+.2f} pts")
        if ensembles:
            be = ensembles[0]
            print(f"    Best ensemble: {be['method']}  acc={be['acc']*100:.2f}%  "
                  f"F1={be['f1']*100:.2f}%  Δ vs best paper: {(be['acc']-best_pub)*100:+.2f} pts")

        all_results[ds_code] = (models, ensembles, papers)

    if not all_results:
        print("No data found. Run 03_01 + 03_03 for KFR and/or MFR fruit_state first.")
        return

    wb = Workbook()
    write_summary(wb.active, all_results)
    for ds_code, (models, ensembles, papers) in all_results.items():
        ws = wb.create_sheet(f"{ds_code} vs Papers")
        write_ds_sheet(ws, ds_code, models, ensembles, papers)

    wb.save(OUT)
    sheets = ", ".join(wb.sheetnames)
    print(f"\n  Saved: {OUT}  ({len(wb.sheetnames)} sheets: {sheets})")


if __name__ == "__main__":
    main()
