"""
generate_tracker.py — Generate experiments_tracker.xlsx from existing run_outputs/.

Structure:
  One sheet per backbone + one SUMMARY sheet.
  Rows: dataset x label_mode combinations.
  Columns: C1, C2, C3, C4.
  Cell values: "XX.X% F1" if complete, "pending", "partial", "n/a".

Usage:
    python generate_tracker.py
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────
RUN_DIR  = Path("run_outputs")
OUT_FILE = Path("experiments_tracker.xlsx")

BACKBONES = [
    "resnet18",
    "resnet34",
    "resnet50",
    "mobilenet_v3_small",
    "efficientnet_b0",
    "efficientnet_b2",
]

# Ordered smallest to largest (by image count)
DATASETS = [
    "kaggle_fruits_quality",      #    359 imgs
    "mendeley_fruits",            #  1,655 imgs
    "mendeley_lemon_varieties",   #  1,956 imgs
    "mendeley_fruitvision",       # 10,154 imgs
    "kaggle_fruits_fresh_rotten", # 13,599 imgs
    "kaggle_fresh_stale",         # 27,317 imgs
]

DATASET_ORDER = {ds: i+1 for i, ds in enumerate(DATASETS)}

# datasets where fruit_state doesn't apply
FRUIT_STATE_NA = {
    "mendeley_lemon_varieties",  # single fruit
    "kaggle_fruits_quality",     # flat binary dataset
}

CONDITIONS   = ["frozen", "layer4", "head_frozen", "head_layer4"]
COND_LABELS  = ["C1 frozen", "C2 layer4", "C3 head_frozen", "C4 head_layer4"]
LABEL_MODES  = ["state", "fruit_state"]

# ── Short ID mappings ─────────────────────────────────────────
DS_CODE = {
    "kaggle_fruits_quality"     : "KFQ",
    "mendeley_fruits"           : "MFR",
    "mendeley_lemon_varieties"  : "MLM",
    "mendeley_fruitvision"      : "MFV",
    "kaggle_fruits_fresh_rotten": "KFR",
    "kaggle_fresh_stale"        : "KFS",
    "own_dataset"               : "OWN",
}
BB_CODE = {
    "resnet18"          : "R18",
    "resnet34"          : "R34",
    "resnet50"          : "R50",
    "mobilenet_v3_small": "MN3",
    "efficientnet_b0"   : "EB0",
    "efficientnet_b2"   : "EB2",
}
COND_CODE = {
    "frozen"     : "C1",
    "layer4"     : "C2",
    "head_frozen": "C3",
    "head_layer4": "C4",
}
MODE_CODE = {
    "state"      : "ST",
    "fruit_state": "FS",
}

def make_id(dataset, backbone, condition, label_mode):
    ds = DS_CODE.get(dataset, dataset[:3].upper())
    bb = BB_CODE.get(backbone, backbone[:3].upper())
    cc = COND_CODE.get(condition, condition)
    lm = MODE_CODE.get(label_mode, label_mode[:2].upper())
    return f"{ds}-{bb}-{cc}-{lm}"

# ── Colors ────────────────────────────────────────────────────
COLOR_COMPLETE  = "C6EFCE"   # green
COLOR_PARTIAL   = "FFEB9C"   # yellow
COLOR_PENDING   = "FFFFFF"   # white
COLOR_NA        = "D9D9D9"   # grey
COLOR_HEADER    = "1F4E79"   # dark blue
COLOR_SUBHEADER = "2E75B6"   # medium blue
COLOR_DS_ROW    = "DEEAF1"   # light blue

# ── Load all metrics ──────────────────────────────────────────

def load_runs():
    """
    Return dict keyed by (dataset, condition, backbone, label_mode) → metrics.
    Only counts runs with new metrics format (has best_mcc).
    Old format runs and incomplete runs are ignored (treated as pending).
    Runs with checkpoint but no metrics.json are marked as 'running'.
    """
    runs = {}
    for run_dir in RUN_DIR.iterdir():
        if not run_dir.is_dir() or run_dir.name == "_plots":
            continue
        mpath = run_dir / "metrics.json"
        cpath = run_dir / "checkpoint.pt"

        if not mpath.exists():
            # Has checkpoint but no metrics = currently running
            if cpath.exists():
                runs[(run_dir.name, None, None, None)] = {"status": "running"}
            continue

        with open(mpath) as f:
            m = json.load(f)

        # Skip old format runs (no best_mcc = needs rerun)
        if "best_mcc" not in m:
            continue

        key = (
            m.get("dataset", ""),
            m.get("condition", ""),
            m.get("backbone_name", "resnet18"),
            m.get("label_mode", "state") or "state",
        )
        evals = len(m.get("acc_history", []))
        f1    = m.get("best_f1", 0)
        runs[key] = {
            "status"   : "complete" if evals >= 10 else "partial",
            "best_f1"  : f1,
            "best_acc" : m.get("best_acc", 0),
            "best_mcc" : m.get("best_mcc", None),
            "best_auc" : m.get("best_auc", None),
            "evals"    : evals,
            "time_min" : round(m.get("training_time_seconds", 0) / 60, 1),
        }
    return runs


def is_running(runs, dataset, condition, backbone, label_mode):
    """Check if a run is currently in progress (checkpoint exists, no metrics)."""
    # Match by run folder name pattern
    mode_suffix = "_fs" if label_mode == "fruit_state" else ""
    run_name = f"{dataset}_all_{condition}_{backbone}{mode_suffix}"
    return (run_name, None, None, None) in runs


def cell_value(runs, dataset, condition, backbone, label_mode):
    """Return display string and status for a cell."""
    # n/a cases
    if label_mode == "fruit_state" and dataset in FRUIT_STATE_NA:
        return "n/a", "na"

    key = (dataset, condition, backbone, label_mode)
    if key not in runs:
        # Check if currently running
        if is_running(runs, dataset, condition, backbone, label_mode):
            return "running", "running"
        return "pending", "pending"

    r = runs[key]
    if r["status"] == "partial":
        return "partial", "partial"

    f1  = r["best_f1"]
    acc = r["best_acc"]
    mcc = r["best_mcc"]
    auc = r["best_auc"]
    t   = r["time_min"]

    mcc_str = f"{mcc:.3f}" if mcc is not None else "n/a"
    auc_str = f"{auc:.4f}" if auc is not None else "n/a"

    run_id = make_id(dataset, backbone, condition, label_mode)

    text = (f"{run_id}\n"
            f"F1:  {f1*100:.1f}%\n"
            f"ACC: {acc*100:.1f}%\n"
            f"MCC: {mcc_str}\n"
            f"AUC: {auc_str}\n"
            f"Time: {t} min")
    return text, "complete"


# ── Style helpers ─────────────────────────────────────────────

def auto_fit_columns(ws, min_width=8, max_width=60):
    """Set column widths based on the longest content in each column."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                lines = str(cell.value).split("\n")
                max_len = max(max_len, max(len(line) for line in lines))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def style_cell(cell, value, status, bold=False):
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()
    if status == "complete":
        cell.fill = make_fill(COLOR_COMPLETE)
    elif status == "partial":
        cell.fill = make_fill(COLOR_PARTIAL)
    elif status == "na":
        cell.fill = make_fill(COLOR_NA)
        cell.font = Font(color="888888", italic=True)
    else:
        cell.fill = make_fill(COLOR_PENDING)
    if bold:
        cell.font = Font(bold=True)


def write_backbone_sheet(ws, backbone, runs):
    """Write one backbone sheet — minimal, no colors."""

    # ── Header row ──
    headers = ["Order", "Dataset", "Label Mode", "C1 frozen", "C2 layer4",
               "C3 head_frozen", "C4 head_layer4"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    # ── Data rows ──
    row = 2
    for ds in DATASETS:
        order = DATASET_ORDER[ds]
        for lm in LABEL_MODES:
            ws.cell(row=row, column=1, value=order)
            ws.cell(row=row, column=2, value=ds)
            ws.cell(row=row, column=3, value=lm)

            for col, cond in enumerate(CONDITIONS, 4):
                val, _ = cell_value(runs, ds, cond, backbone, lm)
                c = ws.cell(row=row, column=col, value=val)
                c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            row += 1

    # ── Column widths ──
    auto_fit_columns(ws)


def write_summary_sheet(ws, runs):
    """Write summary sheet — best F1 per dataset across all backbones (C4, state)."""

    # Header
    headers = ["Order", "Dataset"] + BACKBONES
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    # Data
    for row, ds in enumerate(DATASETS, 2):
        ws.cell(row=row, column=1, value=DATASET_ORDER[ds])
        ws.cell(row=row, column=2, value=ds)

        for col, bb in enumerate(BACKBONES, 3):
            val, _ = cell_value(runs, ds, "head_layer4", bb, "state")
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="center")

    auto_fit_columns(ws)


# ── Legend sheet ──────────────────────────────────────────────

def write_legend_sheet(ws):
    row = 1

    def section(title):
        nonlocal row
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

    def entry(col1, col2):
        nonlocal row
        ws.cell(row=row, column=1, value=col1)
        ws.cell(row=row, column=2, value=col2)
        row += 1

    def blank():
        nonlocal row
        row += 1

    # ── Cell status values ──
    section("CELL STATUS VALUES")
    entry("Value",       "Meaning")
    entry("XX.X% F1",   "Experiment complete — cell shows ID + metrics")
    entry("running",    "Experiment currently in progress")
    entry("partial",    "Experiment started but interrupted (no full metrics)")
    entry("pending",    "Experiment not yet run")
    entry("n/a",        "Not applicable for this dataset / label mode combination")

    blank()

    # ── Cell content format ──
    section("CELL CONTENT FORMAT (completed experiments)")
    entry("Line 1",  "ID — short identifier (e.g. KFQ-R18-C2-ST)")
    entry("Line 2",  "F1: macro-averaged F1 score")
    entry("Line 3",  "ACC: accuracy")
    entry("Line 4",  "MCC: Matthews Correlation Coefficient (-1 worst / +1 perfect)")
    entry("Line 5",  "AUC: Area Under ROC Curve (0.5 random / 1.0 perfect)")
    entry("Line 6",  "Time: total training time in minutes")

    blank()

    # ── ID format ──
    section("ID FORMAT:  {DATASET}-{BACKBONE}-{CONDITION}-{LABELMODE}")
    entry("Example", "KFQ-R18-C4-ST  =  kaggle_fruits_quality, ResNet18, head_layer4, state")

    blank()

    # ── Dataset codes ──
    section("DATASET CODES")
    for ds, code in DS_CODE.items():
        entry(code, ds)

    blank()

    # ── Backbone codes ──
    section("BACKBONE CODES")
    for bb, code in BB_CODE.items():
        entry(code, bb)

    blank()

    # ── Condition codes ──
    section("CONDITION CODES")
    entry("C1", "frozen      — backbone frozen + linear head only")
    entry("C2", "layer4      — last backbone block unfrozen + linear head")
    entry("C3", "head_frozen — backbone frozen + projection head (512→256→128) + linear")
    entry("C4", "head_layer4 — layer4 unfrozen + projection head + linear (best results)")

    blank()

    # ── Label mode codes ──
    section("LABEL MODE CODES")
    entry("ST", "state       — classify by freshness state only (fresh / rotten / formalin)")
    entry("FS", "fruit_state — classify by fruit + state (apple_fresh / apple_rotten / ...)")

    blank()

    # ── Sheet structure ──
    section("SHEET STRUCTURE")
    entry("resnet18 … efficientnet_b2", "One sheet per backbone. Rows = dataset × label mode. Columns = C1-C4.")
    entry("SUMMARY",  "Best F1 for C4 head_layer4 state across all backbones and datasets.")
    entry("Legend",   "This sheet.")

    blank()

    # ── Backbone notes ──
    section("BACKBONE NOTES")
    entry("R18 (ResNet-18)",       "Primary baseline. 11.2M params, 512-dim output. All experiments complete.")
    entry("EB0 (EfficientNet-B0)", "Phase 2 backbone. 5.3M params, 1280-dim output. All experiments complete.")
    entry("EB2 (EfficientNet-B2)", "Phase 2 backbone. 9.1M params, 1408-dim output. All experiments complete.")
    entry("MN3 (MobileNetV3-S)",   "Lightweight backbone for edge/mobile deployment. 2.5M params, 576-dim output. Experiments in progress.")
    entry("R34 (ResNet-34)",       "NOT CURRENTLY EVALUATED. Included in registry for future use. "
                                   "Similar architecture to R18 with more layers (21.3M params). "
                                   "Expected to perform marginally better than R18 but at higher computational cost.")
    entry("R50 (ResNet-50)",       "NOT CURRENTLY EVALUATED. Included in registry for future use. "
                                   "Deeper ResNet variant (25.6M params, 2048-dim output). "
                                   "Uses bottleneck blocks; better suited for large datasets.")

    auto_fit_columns(ws)


# ── Main ──────────────────────────────────────────────────────

def main():
    runs = load_runs()
    print(f"Loaded {len(runs)} run records from {RUN_DIR}/")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # One sheet per backbone
    for bb in BACKBONES:
        ws = wb.create_sheet(title=bb)
        write_backbone_sheet(ws, bb, runs)
        print(f"  Sheet: {bb}")

    # Summary sheet
    ws_sum = wb.create_sheet(title="SUMMARY")
    write_summary_sheet(ws_sum, runs)
    print(f"  Sheet: SUMMARY")

    # Legend sheet
    ws_leg = wb.create_sheet(title="Legend")
    write_legend_sheet(ws_leg)
    print(f"  Sheet: Legend")

    wb.save(OUT_FILE)
    print(f"\nSaved: {OUT_FILE}")


if __name__ == "__main__":
    main()
